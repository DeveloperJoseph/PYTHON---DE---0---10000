from openpyxl import Workbook
#new book xls
wb = Workbook()
#grab the active worksheet
ws = wb.active

ws['A1']= 'FILE CREATED ON:'
import datetime
ws['B1']=datetime.datetime.now()

ws['A2']= 'VALUE OF PI'
import math
valuePI = math.pi
ws['B2']= valuePI

ws['A3']='Days:'
ws.append(['Monday','Tuesday','Sunday'])

#Python types will automatically be converted

#Save the file
wb.save("myFirstOpenPyXls.xlsx")
print("> File Save")

