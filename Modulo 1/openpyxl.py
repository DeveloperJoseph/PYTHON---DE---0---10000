
from openpyxl import Workbook

wb = Workbook()

"""
    #new book xls
    
    #grab the active worksheet
    ws = wb.active
    ws['A1']= 'FILE CREATED ON:'
    import datetime
    ws['B1']=datetime.datetime.now()
    ws['B2']=datetime.datetime.year

    ws['A2']= 'VALUE OF PI'
    import math
    valuePI = math.pi
    ws['B2']= valuePI

    ws['A5']='Days:'
    ws.append(['Monday','Tuesday','Sunday'])

    ws['A4']= 'RAFAEL'
    ws['B4']= '12/07/1991'

    #Python types will automatically be converted

    #Save the file
    wb.save("myFirstOpenPyXls.xlsx")
    print("> File Save")


except:
    print(">> Error al importar la libreria")
"""